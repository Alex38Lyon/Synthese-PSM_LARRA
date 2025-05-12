# -*- coding: utf-8 -*-

########################################################################################################################################
#                                                        	                                                                           #  
#                                Script exporter une base de donnée therion vers un BD type Karsteau                                   #
#                                                                                                                                      #                                           
#                                         By Alexandre PONT  alexandre.pont@yahoo.fr                                                   # 
#                                                                                                                                      #
# Commande : python pyThtoBD.py --help                                                                                                 #
# Utilisation:                                                                                                                         #
#   1 : Placer des fichiers Export_bd.ini dans chacun des dossiers des cavités à exporter                                              #
#   2 : Lancer python pyThtoBD.py, sélectionner le dossier therion à exporter                                                          # 
#   3 : Résultats pour Karsteau dans le dossier /Outputs/Export_bd/                                                                    #    
#   4 : (A venir - Importer le résultat de l'importation dans Karsteau )                                                               #   
#                                                                                                                                      #   
#                                                                                                                                      #   
########################################################################################################################################

'''
To do list :

- A gérer le cas de données supprimées dans th... les supprimer de la BD
- Fonction pour importer les clés Karsteau

'''

import sqlite3, sys, os, re, argparse
import pandas as pd
import configparser
import hashlib
import tkinter as tk
import zipfile
from tkinter import filedialog
from alive_progress import alive_bar              # https://github.com/rsalmei/alive-progress	
from datetime import datetime
from os.path import abspath
from Lib.therion import Colors, compile_file, get_stats_from_log, get_syscoord_from_log
from Lib.logger_config import setup_logger
from PyPDF2 import PdfReader
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

Version ="2025.05.12"  

export_file ="Export_bd.ini"
export_folder = "/Outputs/Export_bd/"
export_db = "Export_bd.db"
therion_path = "C:/Program Files/Therion/therion.exe"
log_file = "Export_bd.log"
debug_log = False  


#####################################################################################################################################
#         Fonction calculer le hash d'un fichier                                                                                    #
#####################################################################################################################################
def hash_file(filepath, algo='sha256', chunk_size=8192):
    hasher = hashlib.new(algo)
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(chunk_size), b''):
            hasher.update(chunk)
    return hasher.hexdigest()


#####################################################################################################################################
#         Fonction pour importer un fichier SQL dans une base de données SQLite                                                     #
#####################################################################################################################################
def importation_sql_db(fichier_sql, fichier_db):
    """
    Fonction pour importer un fichier SQL dans une base de données SQLite

    Args:
        fichier_sql (_type_): _description_
    """         
         
    global error_count
    
    try:
        # Si la base de données existe, supprimez-la pour forcer l'écriture
        
        if os.path.exists(fichier_db):
            #print("Suppression de la Bd existante: " + imported_database)
            os.remove(fichier_db)
            
        connection = sqlite3.connect(fichier_db)
        cursor = connection.cursor()
        
        # Lecture du fichier SQL et exécution des commandes
        with open(fichier_sql, 'r') as sql_file:
            sql_script = sql_file.read()
             
        
        # Séparation du script en commandes individuelles
        sql_script = re.sub(r', nan', ', 0', sql_script, flags=re.IGNORECASE)
        commandes = [cmd.strip() + ';\n' for cmd in sql_script.split(';\n') if cmd.strip()]
       

        # Exécution des commandes avec une barre de progression
        with alive_bar(len(commandes), title = f"{Colors.GREEN}\tImportation base sql : {Colors.WHITE}{os.path.basename(fichier_sql)}",  length = 20) as bar:
            for commande in commandes:
                cursor.execute(commande)
                connection.commit()
                bar()

        connection.close()
        
        log.info(f"Importation réussie de la base de données Therion : {Colors.MAGENTA}{fichier_db}")
        
    except sqlite3.Error as e:
        log.error(f"Erreur lors de l'exécution de la requête importation_sql_data code:{Colors.CYAN}{e}") 
        error_count  += 1
        # sys.exit(1)  # Arrête le programme en cas d'erreur
        
    return


#####################################################################################################################################
#        Requête : Table des entrées  (Liste des entrées avec coordonnées)                                                          #
#####################################################################################################################################
def sql_liste_entree(cursor, ID_CAVITE, NAME):
    global error_count, warning_fix
    
    
    """
    Retour une table avec en ligne :
        ID Station
        Name Station
        X
        Y
        Z
    """
         
    sql_query_ent=  ("""
                    select 
                        STATION.ID, 
                        STATION.NAME, 
                        /*SURVEY.NAME, SURVEY.PARENT_ID, SURVEY.FULL_NAME, SURVEY.TITLE,*/ 
                        round(STATION.X, 1), 
                        round(STATION.Y, 1), 
                        round(STATION.Z, 1),
                        SURVEY.NAME,
                        SURVEY.TITLE
                        /*, STATION_FLAG.FLAG , count(STATION.NAME) AS Nombre_Occurrences */
                    from STATION 
                    join STATION_FLAG on STATION_FLAG.STATION_ID = STATION.ID
                    join SURVEY on SURVEY.ID = STATION.SURVEY_ID
                    where  STATION_FLAG.FLAG='ent' -- or STATION_FLAG.FLAG='fix' --and STATION.ID = 28548 
                    group by STATION.NAME , STATION.Y, STATION.Z 
                    order by STATION.NAME ASC 
                """)
    
    sql_query_fix=  ("""
                        SELECT 
                            STATION.ID, 
                            STATION.NAME, 
                            ROUND(STATION.X, 1), 
                            ROUND(STATION.Y, 1), 
                            ROUND(STATION.Z, 1),
                            SURVEY.NAME,
                            SURVEY.TITLE
                        FROM STATION
                        JOIN STATION_FLAG ON STATION_FLAG.STATION_ID = STATION.ID
                        JOIN SURVEY ON SURVEY.ID = STATION.SURVEY_ID
                        WHERE STATION.ID IN (
                            SELECT STATION_ID
                            FROM STATION_FLAG
                            WHERE FLAG = 'fix'
                        )
                        AND STATION.ID NOT IN (
                            SELECT STATION_ID
                            FROM STATION_FLAG
                            WHERE FLAG = 'ent'
                        )
                        GROUP BY STATION.ID, STATION.NAME, STATION.X, STATION.Y, STATION.Z, SURVEY.NAME, SURVEY.TITLE
                        ORDER BY STATION.NAME ASC;
            """)
    
    try:
        
        cursor.execute(sql_query_fix)    
        result_fix = cursor.fetchall()
        
        if len(result_fix) > 0 :
            log.warning(f"{Colors.CYAN}{len(result_fix)}{Colors.YELLOW} point(s) fixe trouvées pour la cavité ID : {Colors.CYAN}[{ID_CAVITE}] {Colors.WHITE}{NAME}")
            warning_fix += len(result_fix)
                
        cursor.execute(sql_query_ent)    
        result_ent = cursor.fetchall()

        if len(result_ent) == 0 :
            error_count  += 1 
            log.error(f"\tAttention aucune entrée ou point fix trouvé")
        else :
            log.debug(f"\t Table des STATION, entrée nbre: {Colors.MAGENTA}{len(result_ent)}")
        
        return result_ent
        
    except sqlite3.Error as e:
        log.error(f"Erreur lors de l'exécution de la requête 4 (sql_liste_entree) : {Colors.CYAN}{e}")
        error_count  += 1
        return None


#####################################################################################################################################
#            Création BD                                                                                                            #
#####################################################################################################################################      
def create_new_db(db_path):
    """
    Crée une table SQLite nommée `table_name` avec les champs spécifiés.
    
    :param db_path: Chemin du fichier .sqlite (sera créé s'il n'existe pas)
    :param table_name: Nom de la table à créer
    :return: Objet de connexion sqlite3.Connection
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    create_table_CAVITE = f'''
    CREATE TABLE IF NOT EXISTS CAVITE (
        ID INTEGER PRIMARY KEY AUTOINCREMENT,
        NAME TEXT CHECK(LENGTH(NAME) <= 64),
        SYNO_1 TEXT CHECK(LENGTH(SYNO_1) <= 64),
        SYNO_2 TEXT CHECK(LENGTH(SYNO_2) <= 64),
        SYNO_3 TEXT CHECK(LENGTH(SYNO_3) <= 64),
        DEV REAL,
        DENIV_PLUS REAL,
        DENIV_MOINS REAL,
        PATH TEXT CHECK(LENGTH(PATH) <= 128),
        TH_VALIDE BOOLEAN DEFAULT FALSE,
        KEY_KARSTEAU INTEGER,
        HASH_SQL_FILE,
        DATE_UPDATE -- format attendu : 'YYYY-MM-DD HH:MM:SS'
    );
    '''
    cursor.execute(create_table_CAVITE)
    
    create_table_ENTREE = f'''
    CREATE TABLE IF NOT EXISTS ENTREE (
        ID INTEGER PRIMARY KEY AUTOINCREMENT,
        ID_CAVITE INTEGER,
        NUM INTEGER,
        NAME TEXT CHECK(LENGTH(NAME) <= 64),
        SYNO_1 TEXT CHECK(LENGTH(SYNO_1) <= 64),
        SYNO_2 TEXT CHECK(LENGTH(SYNO_2) <= 64),
        SYNO_3 TEXT CHECK(LENGTH(SYNO_3) <= 64),
        MARQ_1 TEXT CHECK(LENGTH(MARQ_1) <= 64),
        MARQ_2 TEXT CHECK(LENGTH(MARQ_2) <= 64),
        MARQ_3 TEXT CHECK(LENGTH(MARQ_3) <= 64),
        COORD_X REAL,
        COORD_Y REAL,
        COORD_Z REAL,
        UNIT_COORD TEXT DEFAULT 'm' CHECK(LENGTH(UNIT_COORD) <= 8),
        SYS_COORD TEXT CHECK(LENGTH(SYS_COORD) <= 8),
        ZONE_COORD TEXT CHECK(LENGTH(ZONE_COORD) <= 8),
        METHODE_COORD TEXT CHECK(LENGTH(METHODE_COORD) <= 64),
        SOURCE_COORD TEXT DEFAULT 'Therion topo' CHECK(LENGTH(SOURCE_COORD) <= 64),
        DATE_COORD TEXT CHECK(LENGTH(DATE_COORD) <= 64),
        ACCES_LIBRE TEXT DEFAULT 'N' CHECK(LENGTH(ACCES_LIBRE) <= 1),
        -- Manque les typologies d'entrées --
        PATH TEXT CHECK(LENGTH(PATH) <= 128),
        TH_VALIDE BOOLEAN DEFAULT FALSE,
        KEY_KARSTEAU INTEGER,
        DATE_UPDATE -- format attendu : 'YYYY-MM-DD HH:MM:SS'
    );
    '''
    cursor.execute(create_table_ENTREE)
    
    create_table_DOCUMENT = f'''
    CREATE TABLE IF NOT EXISTS DOCUMENT (
        ID INTEGER PRIMARY KEY AUTOINCREMENT,
        ID_CAVITE INTEGER,
        LIE TEXT CHECK(LENGTH(LIE) <= 64),
        NATURE TEXT CHECK(LENGTH(NATURE) <= 64),
        TITRE TEXT CHECK(LENGTH(TITRE) <= 64),
        DATE TEXT CHECK(LENGTH(DATE) <= 64),
        AUTEUR TEXT CHECK(LENGTH(AUTEUR) <= 64),
        DESCRIPTION TEXT CHECK(LENGTH(DESCRIPTION) <= 128),
        TEXT TEXT CHECK(LENGTH(TEXT) <= 128),
        CAT INTERGER,
        FILE TEXT CHECK(LENGTH(FILE) <= 128),
        HASH_FILE TEXT CHECK(LENGTH(HASH_FILE) <= 128),
        TH_VALIDE BOOLEAN DEFAULT FALSE, 
        KEY_KARSTEAU INTEGER,
        DATE_UPDATE -- format attendu : 'YYYY-MM-DD HH:MM:SS'
    );
    '''
    cursor.execute(create_table_DOCUMENT)
    conn.commit()
        
    return conn     


#####################################################################################################################################
#            Initialisation de la BD  (pour la détection des données supprimées dans Therion                                        #                                                                                   #
#####################################################################################################################################  
def init_th_valide(conn):
    cursor = conn.cursor()      
    cursor.execute("UPDATE CAVITE SET TH_VALIDE = FALSE")
    cursor.execute("UPDATE ENTREE SET TH_VALIDE = FALSE")
    cursor.execute("UPDATE DOCUMENT SET TH_VALIDE = FALSE")
    conn.commit()   
    return

        
#####################################################################################################################################
#            Nom de la cavité                                                                                                       #
#####################################################################################################################################      
# extrait du champs -title du fichier -tot.th
def cave_name(_path_name):  
    global error_count

    tot_files = []
    for file_name in os.listdir(_path_name):
        if file_name.endswith("-tot.th") and os.path.isfile(os.path.join(_path_name, file_name)):
            tot_files.append(os.path.join(_path_name, file_name))
    
    if len(tot_files) > 1 :
        log.error(f"Erreur plusieurs fichiers -tot.th : {Colors.MAGENTA}{tot_files}")
        error_count += 1
        return 
    elif  tot_files == [] :
        log.error(f"Erreur pas de fichiers -tot.th : {Colors.MAGENTA}{_path_name}")
        error_count += 1
        return
    
    with open(tot_files[0], 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip().startswith("survey") and "-title" in line:
                parts = line.split("-title", 1)
                if len(parts) > 1:
                    # On enlève les éventuels guillemets et espaces autour
                    caveName = parts[1].strip().strip('"')
        
    return caveName


#####################################################################################################################################
#            Execution du fichier thconfig                                                                                          #
#####################################################################################################################################      
def exe_therion_cave(_path_name):
    global error_count
    
    ligne_export_db = "export database "
    
    tot_files = []
    for file_name in os.listdir(_path_name):
        if file_name.endswith(".thconfig") and os.path.isfile(os.path.join(_path_name, file_name)):
            tot_files.append(os.path.join(_path_name, file_name))
    
    if len(tot_files) > 1 :
        log.error(f"Erreur plusieurs fichiers .thconfig : {Colors.MAGENTA}{tot_files}")
        error_count += 1
        return
    elif  tot_files == [] :
        log.error(f"Erreur pas de fichiers .thconfig : {Colors.MAGENTA}{_path_name}")
        error_count += 1
        return 
    
    with open(tot_files[0], 'r+', encoding='utf-8') as f:
        lignes = f.readlines()
        # Cherche une ligne commençant par la commande
        ligne_existante = any(l.strip().startswith(ligne_export_db) for l in lignes)

        if not ligne_existante:
            f.write("\n\n# Ajout automatique by script pyThtoBF.py pour export\n")    
            f.write(ligne_export_db + '-o Outputs/database_export_db.sql\n')
            log.debug(f"Ajout automatique de la commande : {Colors.MAGENTA}{ligne_export_db}-o Outputs/database_export_db.sql {Colors.YELLOW}dans le fichier : {Colors.MAGENTA}{tot_files[0]}")
    
    # print(tot_files[0])
        
    compile_file(tot_files[0].replace("\\", "/"))                           
        
        
#####################################################################################################################################
#            Développement et profondeur de la cavité                                                                               #
#####################################################################################################################################      
def cave_dev(_path_name):  
    """
    Extrait les données de développement et dénivelé depuis le fichier log de therion 
    
    """
    global error_count
    
    tot_files = []
    for file_name in os.listdir(_path_name):
        if file_name.endswith("therion.log") and os.path.isfile(os.path.join(_path_name, file_name)):
            tot_files.append(os.path.join(_path_name, file_name))
    
    if len(tot_files) > 1 :
        log.error(f"Erreur plusieurs fichiers therion.log : {Colors.MAGENTA}{tot_files}")
        error_count += 1
        return
    elif  tot_files == [] :
        log.error(f"Erreur pas de fichiers therion.log : {Colors.MAGENTA}{_path_name}")
        error_count += 1
        return
    
    # print("fichiers log.log" + tot_files[0])
    with open(tot_files[0], "r") as f:
        log_content = f.read()
    # print(log_content)
    stats = get_stats_from_log(log_content)
    return stats


#####################################################################################################################################
#            Système de coordonnées utilisé                                                                                         #
#####################################################################################################################################      
def cave_sys(_path_name):  
    """
    Extrait les données de coordonnées log de therion 
    
    """
    global error_count
    
    tot_files = []
    for file_name in os.listdir(_path_name):
        if file_name.endswith("therion.log") and os.path.isfile(os.path.join(_path_name, file_name)):
            tot_files.append(os.path.join(_path_name, file_name))
    
    if len(tot_files) > 1 :
        log.error(f"Erreur plusieurs fichiers therion.log : {Colors.MAGENTA}{tot_files}")
        error_count += 1
        return None
    elif  tot_files == [] :
        log.error(f"Erreur pas de fichiers therion.log : {Colors.MAGENTA}{_path_name}")
        error_count += 1
        return None
    
    # print("fichiers log.log" + tot_files[0])
    with open(tot_files[0], "r") as f:
        log_content = f.read()
    # print(log_content)
    stats = get_syscoord_from_log(log_content)
    return stats


#####################################################################################################################################
#            Mise à jour pdf à exporter                                                                                             #
#####################################################################################################################################  
def pdf_exif(file):
    """
    Extrait les métadonnées d'un fichier PDF : auteur, titre, sujet, date de création.

    :param file: Chemin du fichier PDF
    :return: Tuple (auteur, titre, sujet, date_creation) avec date au format 'YYYY-MM-DD HH:MM:SS'
    """
    global error_count
    
    try:
        reader = PdfReader(file)
        info = reader.metadata

        author = info.get('/Author')
        title = info.get('/Title')
        subject = info.get('/Subject')
        raw_date = info.get('/CreationDate')

        # Formatage de la date
        date_creation = None
        if raw_date and raw_date.startswith("D:"):
            try:
                # Extrait les éléments (ignore le fuseau horaire pour simplifier)
                date_creation = datetime.strptime(raw_date[2:16], "%Y%m%d%H%M%S")
                date_creation = date_creation.strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                date_creation = raw_date  # En cas d'erreur, renvoyer la chaine brute

        return author, title, subject, date_creation

    except Exception as e:
        log.error(f"Erreur lors de la lecture exif du fichier pdf : {Colors.MAGENTA}{file} {Colors.ERROR}Code : {Colors.MAGENTA}{e}")
        error_count += 1
        return None, None, None, None


#####################################################################################################################################
#            Mise à jour pdf à exporter                                                                                             #
#####################################################################################################################################  
def pdf_update(conn, base_path, path_name, _update, ID_CAVITE, NAME):

    _path_name = path_name + "/Outputs"
    _Nature = "pdf"
    _Lie = "cavité"
    _Cat = "73"
    cursor = conn.cursor()
    
    tot_files = []
    for file_name in os.listdir(_path_name):
        if file_name.endswith(".pdf") and os.path.isfile(os.path.join(_path_name, file_name)):
            tot_files.append(os.path.join(_path_name, file_name))
    
    if len(tot_files) > 1 :
        log.debug(f"Cavité: {Colors.WHITE}{NAME}{Colors.GREEN}, fichiers pdf à exporter : {Colors.WHITE}{len(tot_files)}")
    elif  tot_files == [] :
        log.warning(f"Attention cavité: {Colors.WHITE}{NAME}{Colors.YELLOW}, pas de fichiers pdf : {Colors.MAGENTA}{_path_name}")
        return 

    for file in tot_files:
        _file = os.path.relpath(file, base_path).replace("\\", "/")
        
        _hash_file = hash_file(file)
         
        cursor.execute("SELECT HASH_FILE FROM DOCUMENT WHERE ID_CAVITE = ? AND FILE = ?", (ID_CAVITE, _file,))
        _document = cursor.fetchone()
        
        auteur, titre, sujet, date  = pdf_exif(file.replace("\\", "/"))
       
            
        if _document is None :     
            cursor.execute("INSERT INTO DOCUMENT (FILE) VALUES (?)", (_file,))
            ID_DOCUMENT = cursor.lastrowid
            cursor.execute("""
                           UPDATE DOCUMENT SET 
                           ID_CAVITE = ?, 
                           LIE = ?, 
                           FILE = ?, 
                           NATURE = ?, 
                           AUTEUR = ?, 
                           TITRE = ?, 
                           DATE = ?, 
                           DESCRIPTION = ?, 
                           DATE_UPDATE = ?, 
                           CAT = ?, 
                           HASH_FILE = ?,
                           TH_VALIDE = TRUE 
                           WHERE ID = ?
                        """,
                    (ID_CAVITE, _Lie,_file, _Nature, auteur, titre, date, sujet, _update, _Cat, _hash_file, ID_DOCUMENT))
            log.info(f"Création du document : {Colors.WHITE}{_file}{Colors.GREEN} de la cavité : {Colors.WHITE}{NAME}{Colors.GREEN}")
        else :
            cursor.execute("SELECT ID, AUTEUR, TITRE, DESCRIPTION, HASH_FILE FROM DOCUMENT WHERE ID_CAVITE = ? AND FILE = ?", (ID_CAVITE, _file,))
            row = cursor.fetchone()
            ID_DOCUMENT_DB, auteur_db, titre_db, sujet_db, hash_file_db = row
            if (auteur != auteur_db) or (titre != titre_db) or (sujet != sujet_db)  or (_hash_file != hash_file_db):
                cursor.execute("""
                    UPDATE DOCUMENT 
                    SET LIE = ?, 
                    AUTEUR = ?, 
                    TITRE = ?, 
                    DATE = ?, 
                    DESCRIPTION = ?, 
                    DATE_UPDATE = ?, 
                    CAT = ?, 
                    HASH_FILE = ?,
                    TH_VALIDE = TRUE
                    WHERE ID = ?
                """, 
                (_Lie, auteur, titre, date, sujet, _update, _Cat, _hash_file, ID_DOCUMENT_DB))
                log.info(f"Mise à jour pdf : {Colors.WHITE}{_file}{Colors.GREEN} de la cavité : {Colors.WHITE}{NAME}{Colors.GREEN}")
            else:
                cursor.execute("UPDATE DOCUMENT SET TH_VALIDE = TRUE WHERE ID = ?", (ID_DOCUMENT_DB,))
                log.info(f"Aucun changement pour le pdf : {Colors.WHITE}{_file}{Colors.GREEN}")


    conn.commit()
    
    return None


#####################################################################################################################################
#            Mise à jour kml à exporter                                                                                             #
#####################################################################################################################################  
def kml_update(conn, base_path, path_name, _update, ID_CAVITE, NAME):
    
    global error_count

    _path_name = path_name + "/Outputs"
    _Nature = "kml"
    _Lie = "cavité"
    _Cat = "87"
    cursor = conn.cursor()
    
    tot_files = []
    for file_name in os.listdir(_path_name):
        if file_name.endswith(".kml") and os.path.isfile(os.path.join(_path_name, file_name)):
            tot_files.append(os.path.join(_path_name, file_name))
    
    if len(tot_files) > 1 :
        log.debug(f"Cavité: {Colors.WHITE}{NAME}{Colors.GREEN}, fichiers kml à exporter : {Colors.WHITE}{len(tot_files)}")
    elif  tot_files == [] :
        log.error(f"Erreur cavité: {Colors.WHITE}{NAME}{Colors.ERROR}, pas de fichiers kml à exporter : {Colors.MAGENTA}{_path_name}")
        error_count += 1
        return

    for file in tot_files:
        _file = os.path.relpath(file, base_path).replace("\\", "/")
        
        _hash_file = hash_file(file)
                 
        cursor.execute("SELECT HASH_FILE FROM DOCUMENT WHERE ID_CAVITE = ? AND FILE = ?", (ID_CAVITE, _file,))
        _document = cursor.fetchone()
        
        auteur = " "
        titre = "Polygonale kml" 
        sujet = f"Polygonale de {NAME} au format google earth"
        date = _update
       
        if _document is None :     
            cursor.execute("INSERT INTO DOCUMENT (FILE) VALUES (?)", (_file,))
            ID_DOCUMENT = cursor.lastrowid
            cursor.execute("""
                        UPDATE DOCUMENT SET 
                        ID_CAVITE = ?, 
                        LIE = ?, 
                        FILE = ?, 
                        NATURE = ?, 
                        AUTEUR = ?, 
                        TITRE = ?, 
                        DATE = ?, 
                        DESCRIPTION = ?, 
                        DATE_UPDATE = ?, 
                        CAT = ?, 
                        HASH_FILE = ?,
                        TH_VALIDE = TRUE 
                        WHERE ID = ?
                        """,
                    (ID_CAVITE, _Lie,_file, _Nature, auteur, titre, date, sujet, _update, _Cat, _hash_file, ID_DOCUMENT))
            log.info(f"Création du document : {Colors.WHITE}{_file}{Colors.GREEN} de la cavité : {Colors.WHITE}{NAME}{Colors.GREEN}")
        else :
            cursor.execute("SELECT ID, AUTEUR, TITRE, DESCRIPTION, HASH_FILE FROM DOCUMENT WHERE ID_CAVITE = ? AND FILE = ?", (ID_CAVITE, _file,))
            row = cursor.fetchone()
            ID_DOCUMENT_DB, auteur_db, titre_db, sujet_db, hash_file_db = row
            
            if (auteur != auteur_db) or (titre != titre_db) or (sujet != sujet_db) or (_hash_file != hash_file_db):
                cursor.execute("""
                    UPDATE DOCUMENT 
                    SET LIE = ?, 
                    AUTEUR = ?, 
                    TITRE = ?, 
                    DATE = ?, 
                    DESCRIPTION = ?, 
                    DATE_UPDATE = ?, 
                    CAT = ?, 
                    HASH_FILE = ?,
                    TH_VALIDE = TRUE
                    WHERE ID = ?
                """, 
                (_Lie, auteur, titre, date, sujet, _update, _Cat, _hash_file, ID_DOCUMENT_DB))
                log.info(f"Mise à jour kml : {Colors.WHITE}{_file}{Colors.GREEN} de la cavité : {Colors.WHITE}{NAME}{Colors.GREEN}")
            else:
                cursor.execute("UPDATE DOCUMENT SET TH_VALIDE = TRUE WHERE ID = ?", (ID_DOCUMENT_DB,))
                log.info(f"Aucun changement pour le kml : {Colors.WHITE}{_file}{Colors.GREEN}")


    conn.commit()
    
    return None


#####################################################################################################################################
#            Création des fichiers zip                                                                                              #
##################################################################################################################################### 
def zip_file(path_source, path_dest, _name_zip):
    
    _name_zip = _name_zip.replace("\\", "/")
    
    name_zip = _name_zip + ".zip"
    
    exclude_ext = ['.zip', 
                   '.pdf', 
                   '.kml', 
                   '.kmz',
                   '.html', 
                   '.sql', 
                   '.plt',
                   '.cav',
                   '.svg',
                   '.bd', 
                   '.log',
                   '.lox', 
                   '.shp', 
                   '.dbf', 
                   '.prj', 
                   '.shx', 
                   '.tro', 
                   '.trox', 
                   '.xlsx',
                   '.db',
                   '.git',
                   '.dat',
                   '.dbf',
                   '.ai',
                   '.png',
                   '.jpg']

    exclude_ext = set(ext.lower() for ext in exclude_ext)
    
    zip_path = os.path.join(path_dest, name_zip).replace("\\", "/")

    files_to_zip = []
    for foldername, subfolders, filenames in os.walk(path_source):
        for filename in filenames:
            ext = os.path.splitext(filename)[1].lower()
            if ext not in exclude_ext:
                file_path = os.path.join(foldername, filename)
                arcname = os.path.relpath(file_path, path_source)
                files_to_zip.append((file_path, arcname))

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        with alive_bar(len(files_to_zip), title = f"{Colors.GREEN}\tCompression fichier Therion: {Colors.WHITE}{os.path.basename(name_zip)} ",  length = 20) as bar:
            for file_path, arcname in files_to_zip:
                zipf.write(file_path, arcname)
                bar()

    log.info(f"Archive des données Therion créée : {Colors.WHITE}{zip_path}{Colors.GREEN}, {Colors.WHITE}{os.path.getsize(zip_path)} {Colors.GREEN}octets")


#####################################################################################################################################
#            Création du fichier zip avec les données à exporter                                                                    #
##################################################################################################################################### 
def zip_data(conn, path_dest, base_path, _name_zip):
    
    global error_count
    
    cursor = conn.cursor()
    
    cursor.execute("""
            SELECT FILE    
            FROM DOCUMENT
        """)
    file_list = cursor.fetchall()
   
    zip_path = os.path.join(path_dest, _name_zip).replace("\\", "/")

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        with alive_bar(len(file_list), title = f"{Colors.GREEN}\tCompression fichier données : {Colors.WHITE}{os.path.basename(_name_zip)} ",  length = 20) as bar:
            for file in file_list:
                _file = os.path.join(base_path, str(file[0])).replace("\\", "/")
                if os.path.isfile(_file):
                    zipf.write(_file, str(file[0]).replace("\\", "/"))
                else:
                    log.error(f"Erreur d'export des fichiers : {Colors.CYAN}{file[0]}{Colors.ERROR} non trouvé ou invalide")
                    error_count += 1
                bar()

    log.info(f"Archive des fichiers exportés créée : {Colors.WHITE}{zip_path}{Colors.GREEN}, {Colors.WHITE}{os.path.getsize(zip_path)} {Colors.GREEN}octets")


#####################################################################################################################################
#            Mise à jour des fichiers zip                                                                                           #
#####################################################################################################################################  
def zip_update(conn, base_path, path_name, _update, ID_CAVITE, Name) :
    
    global error_count
    
    _path_name = path_name + "/Outputs"
    _Nature = "autre"
    _Lie = "cavité"
    _Cat = "72"
    cursor = conn.cursor()
       
    zip_file(path_name, _path_name, Name)
    
    tot_files = []
    for file_name in os.listdir(_path_name):
        if file_name.endswith(".zip") and os.path.isfile(os.path.join(_path_name, file_name)):
            tot_files.append(os.path.join(_path_name, file_name))
    
    if len(tot_files) > 1 :
        log.info(f"{Colors.WHITE}{len(tot_files)}{Colors.GREEN} Fichiers .zip à exporter : {Colors.MAGENTA}{tot_files}")
    elif  tot_files == [] :
        log.error(f"Erreur pas de fichiers .zip : {Colors.MAGENTA}{_path_name}")
        error_count += 1
        return

    for file in tot_files:
        _file = os.path.relpath(file, base_path).replace("\\", "/")
        
        _hash_file = hash_file(file)
        
        cursor.execute("SELECT HASH_FILE FROM DOCUMENT WHERE ID_CAVITE = ? AND FILE = ?", (ID_CAVITE, _file,))
        _document = cursor.fetchone()
        
        auteur = " "
        titre = "Données topo brutes" 
        sujet = f"Topo de {Name} au format Therion"
        date = _update 
       
        if _document is None :     
            cursor.execute("INSERT INTO DOCUMENT (FILE) VALUES (?)", (_file,))
            ID_DOCUMENT = cursor.lastrowid
            cursor.execute("""
                        UPDATE DOCUMENT SET 
                        ID_CAVITE = ?, 
                        LIE = ?, 
                        FILE = ?, 
                        NATURE = ?, 
                        AUTEUR = ?, 
                        TITRE = ?, 
                        DATE = ?, 
                        DESCRIPTION = ?, 
                        DATE_UPDATE = ?, 
                        CAT = ?,
                        HASH_FILE = ?, 
                        TH_VALIDE = TRUE    
                        WHERE ID = ?
                    """,
                    (ID_CAVITE, _Lie,_file, _Nature, auteur, titre, date, sujet, _update, _Cat, _hash_file, ID_DOCUMENT))
            log.info(f"Création du zip : {Colors.WHITE}{_file}{Colors.GREEN} de la cavité : {Colors.WHITE}{Name}{Colors.GREEN}")
        else :
            cursor.execute("SELECT ID, AUTEUR, TITRE, DESCRIPTION, HASH_FILE FROM DOCUMENT WHERE ID_CAVITE = ? AND FILE = ?", (ID_CAVITE, _file,))
            row = cursor.fetchone()
            ID_DOCUMENT_DB, auteur_db, titre_db, sujet_db, hash_file_db = row
            if (auteur != auteur_db) or (titre != titre_db) or (sujet != sujet_db)  or (_hash_file != hash_file_db):
                cursor.execute("""
                    UPDATE DOCUMENT 
                    SET LIE = ?, 
                    AUTEUR = ?, 
                    TITRE = ?, 
                    DATE = ?, 
                    DESCRIPTION = ?, 
                    DATE_UPDATE = ?, 
                    CAT = ?, 
                    HASH_FILE = ?,
                    TH_VALIDE = TRUE
                    WHERE ID = ?
                """, 
                (_Lie, auteur, titre, date, sujet, _update, _Cat, _hash_file, ID_DOCUMENT_DB))
                log.info(f"Mise à jour zip : {Colors.WHITE}{_file}{Colors.GREEN} de la cavité : {Colors.WHITE}{Name}{Colors.GREEN}")
            else:
                cursor.execute("UPDATE DOCUMENT SET TH_VALIDE = TRUE WHERE ID = ?", (ID_DOCUMENT_DB,))
                log.info(f"Aucun changement pour le zip : {Colors.WHITE}{_file}{Colors.GREEN}")


    conn.commit()
    
    
    return None


#####################################################################################################################################
#            Mise à jour des entrées                                                                                                #
#####################################################################################################################################  
def entrance_update(conn, base_path, path_name, _update, ID_CAVITE, NAME, syscoord):
    
    global error_count
    
    _path_name = path_name + "/Outputs"
    
    tot_files = []
    for file_name in os.listdir(_path_name):
        if file_name.endswith(".sql") and os.path.isfile(os.path.join(_path_name, file_name)):
            tot_files.append(os.path.join(_path_name, file_name))
    
    if len(tot_files) > 1 :
        log.error(f"Erreur plusieurs fichiers sql : {Colors.MAGENTA}{tot_files}")
        error_count += 1
        return
    elif  tot_files == [] :
        log.error(f"Erreur pas de fichiers sql : {Colors.MAGENTA}{_path_name}")
        error_count += 1
        return 
    
    cursor = conn.cursor()
    cursor.execute("SELECT HASH_SQL_FILE FROM CAVITE WHERE ID = ?", (ID_CAVITE,))
    _hash_sql = cursor.fetchone()[0]
      
    imported_database = tot_files[0][:-4]+".db"
    sql_file = tot_files[0].replace("\\", "/")
   
    new_hash_file = hash_file(sql_file)
    
    if new_hash_file != _hash_sql :
        cursor.execute("UPDATE CAVITE SET HASH_SQL_FILE = ? WHERE ID = ?", (new_hash_file, ID_CAVITE,))
        importation_sql_db(sql_file, imported_database)
    else : 
        log.info(f"Aucun changement pour la base sql : {Colors.WHITE}{sql_file}{Colors.GREEN} de la cavité : {Colors.WHITE}{NAME}{Colors.GREEN}")
        
    conn2 = sqlite3.connect(imported_database)  # Connexion à la base de données SQLite
    cursor2 = conn2.cursor()
    
    resultat = sql_liste_entree(cursor2, ID_CAVITE, NAME)
    # df = pd.DataFrame(resultat, columns=["ID", "Name", "X", "Y", "Z", "Survey"])
    # print(df)
    
    
    cursor.execute("SELECT COUNT(*) FROM ENTREE WHERE ID_CAVITE = ?", (ID_CAVITE,))
    _entree = cursor.fetchone()[0]
    
    log.info(f"Nombre d'entrées existantes : {Colors.WHITE}{_entree}{Colors.GREEN} trouvées pour la cavité ID : {Colors.CYAN}[{ID_CAVITE}] {Colors.WHITE}{NAME}")

    for line in resultat:
        Entre_Th = line[1] + "@" + line[5]
        cursor.execute("SELECT COUNT(*) FROM ENTREE WHERE ID_CAVITE = ? AND PATH = ?", (ID_CAVITE, Entre_Th,))
        _entree = cursor.fetchone()[0]
        
        if _entree == 0 :
            cursor.execute("INSERT INTO ENTREE (PATH) VALUES (?)", (Entre_Th,))
            ID_ENTREE = cursor.lastrowid
            cursor.execute("""
                        UPDATE ENTREE SET 
                        ID_CAVITE = ?, 
                        NAME = ?, 
                        COORD_X = ?, 
                        COORD_Y = ?, 
                        COORD_Z = ?, 
                        DATE_UPDATE = ?, 
                        SYS_COORD = ?, 
                        TH_VALIDE = TRUE 
                        WHERE ID = ?
                    """,
                    (ID_CAVITE, line[6], line[2], line[3], line[4], _update, syscoord, ID_ENTREE))
            log.info(f"Création de l'entrée : {Colors.WHITE}{Entre_Th}{Colors.GREEN} de la cavité : {Colors.WHITE}{NAME}{Colors.GREEN}")
        elif _entree == 1 :
            cursor.execute("SELECT ID, NAME, COORD_X, COORD_Y, COORD_Z, SYS_COORD FROM ENTREE WHERE ID_CAVITE = ? AND PATH = ?", (ID_CAVITE, Entre_Th,))
            row = cursor.fetchone()
            ID_ENTREE_DB, name_db, x_db, y_db, z_db, syscoord_db = row

            name_new, x_new, y_new, z_new, syscoord_new = line[6], line[2], line[3], line[4], syscoord

            if (name_db != name_new) or (x_db != x_new) or (y_db != y_new) or (z_db != z_new) or (syscoord_db != syscoord_new):
                log.info(f"Mise à jour de l'entrée : {Colors.WHITE}{Entre_Th}{Colors.GREEN} de la cavité : {Colors.WHITE}{NAME}")
                cursor.execute("""
                            UPDATE ENTREE SET 
                            NAME = ?, 
                            COORD_X = ?, 
                            COORD_Y = ?, 
                            COORD_Z = ?, 
                            DATE_UPDATE = ?, 
                            SYS_COORD = ?, 
                            TH_VALIDE = TRUE 
                            WHERE ID = ?
                        """,
                        (name_new, x_new, y_new, z_new, _update, syscoord, ID_ENTREE_DB))
            else:
                cursor.execute("UPDATE ENTREE SET TH_VALIDE = TRUE WHERE ID = ?", (ID_ENTREE_DB,))
                log.info(f"Aucun changement pour l'entrée : {Colors.WHITE}{Entre_Th}{Colors.GREEN} de la cavité : {Colors.WHITE}{NAME}")
            
        else :
            log.error(f"Erreur, entrée en doublons vérifier la base de données:  {Colors.CYAN}{Entre_Th} Nbe {Entre_Th}")
            error_count += 1
    
    conn.commit()
        
    conn2.close()
    
    return None

     
#####################################################################################################################################
#            Update BD                                                                                                              #
#####################################################################################################################################      
def cavite_update(conn, file_list, base_path, _update):
    cursor = conn.cursor()

    for line in file_list:
        ID_CAVITE = ""
        file_path = abspath(line).replace("\\", "/")
        path_name = os.path.dirname(file_path).replace("\\", "/")
        rel_path_name = os.path.relpath(path_name, base_path).replace("\\", "/")
        rel_file_name = os.path.relpath(file_path, base_path).replace("\\", "/")

        exe_therion_cave(path_name)
        cave_Name = cave_name(path_name)
        cave_Dev = cave_dev(path_name)
        cave_Sys = cave_sys(path_name)

        cave_config = configparser.ConfigParser()

        if os.stat(file_path).st_size == 0:
            log.info(f"Le fichier {rel_file_name} est vide. Création avec ID_CAVITE = ''")
            cave_config['Data_Export'] = {'ID_CAVITE': ''}
            with open(file_path, 'w', encoding='utf-8') as configfile:
                cave_config.write(configfile)

        cave_config.read(file_path, encoding="utf-8")

        if 'Data_Export' in cave_config and 'ID_CAVITE' in cave_config['Data_Export']:
            ID_CAVITE = cave_config['Data_Export']['ID_CAVITE']

        insertion_necessaire = False
        if ID_CAVITE == "":
            insertion_necessaire = True
        else:
            cursor.execute("SELECT 1 FROM CAVITE WHERE ID = ?", (ID_CAVITE,))
            exists = cursor.fetchone()
            if not exists:
                insertion_necessaire = True

        if insertion_necessaire:
            cursor.execute("INSERT INTO CAVITE (NAME) VALUES (?)", (cave_Name,))
            ID_CAVITE = cursor.lastrowid
            cursor.execute("""
                        UPDATE CAVITE SET 
                        PATH = ?, 
                        DEV = ?, 
                        DENIV_MOINS = ?, 
                        DENIV_PLUS = ?, 
                        DATE_UPDATE = ?, 
                        TH_VALIDE = TRUE 
                        WHERE ID = ?
                    """,
                    (rel_path_name, float(cave_Dev['length']), float(cave_Dev['depth']), 0.0, _update, ID_CAVITE))
            log.info(f"Nouvelle cavité insérée avec ID : {Colors.MAGENTA}{ID_CAVITE}{Colors.YELLOW}, Name : {Colors.MAGENTA}{cave_Name}{Colors.YELLOW}, Dev : {Colors.MAGENTA}{cave_Dev['length']}m{Colors.YELLOW}, Prof : {Colors.MAGENTA}{cave_Dev['depth']}m{Colors.YELLOW}, Len path : {Colors.MAGENTA}{len(path_name)}")
            cave_config['Data_Export'] = {'ID_CAVITE': str(ID_CAVITE)}
            with open(file_path, 'w', encoding='utf-8') as configfile:
                cave_config.write(configfile)

        else:
            cursor.execute("SELECT NAME, PATH, DEV, DENIV_MOINS, DENIV_PLUS FROM CAVITE WHERE ID = ?", (ID_CAVITE,))
            db_name, db_path, db_dev, db_depth, db_plus = cursor.fetchone()

            new_name = cave_Name
            new_path = rel_path_name
            new_dev = float(cave_Dev['length'])
            new_depth = float(cave_Dev['depth'])
            new_plus = 0.0

            if (db_name != new_name or db_path != new_path or db_dev != new_dev or db_depth != new_depth or db_plus != new_plus):
                log.info(f"Mise à jour cavité ID : {Colors.MAGENTA}{ID_CAVITE}{Colors.YELLOW}, Name : {Colors.MAGENTA}{cave_Name}{Colors.YELLOW}, Dev : {Colors.MAGENTA}{cave_Dev['length']}{Colors.YELLOW}, Prof : {Colors.MAGENTA}{cave_Dev['depth']}{Colors.YELLOW}, Len path : {Colors.MAGENTA}{len(path_name)}")
                cursor.execute("""
                            UPDATE CAVITE SET 
                            NAME = ?, 
                            PATH = ?, 
                            DEV = ?, 
                            DENIV_MOINS = ?, 
                            DENIV_PLUS = ?, 
                            DATE_UPDATE = ?, 
                            TH_VALIDE = TRUE 
                            WHERE ID = ?
                        """,
                        (new_name, new_path, new_dev, new_depth, new_plus, _update, ID_CAVITE))
            else:
                cursor.execute("UPDATE CAVITE SET TH_VALIDE = TRUE WHERE ID = ?", (ID_CAVITE,))
                log.debug(f"Aucune modification pour la cavité ID : {Colors.MAGENTA}{ID_CAVITE} {cave_Name}")

        conn.commit()
        
        entrance_update(conn, base_path, path_name, _update, ID_CAVITE, cave_Name, cave_Sys['syscoord'])
        pdf_update(conn, base_path, path_name, _update, ID_CAVITE, cave_Name)
        kml_update(conn, base_path, path_name, _update, ID_CAVITE, cave_Name)
        zip_update(conn, base_path, path_name, _update, ID_CAVITE, cave_Name)
       

#####################################################################################################################################
#            Recherche fichiers BD                                                                                                  #
#####################################################################################################################################  
def find_files(selected_folder, pattern):
    """
    Recherche récursive des fichiers contenant un motif donné dans leur nom.

    :param selected_folder: Chemin du dossier racine à explorer
    :param pattern: Motif à rechercher dans les noms de fichiers (ex: 'export_BD')
    :return: Liste des chemins complets des fichiers correspondants
    """
    matching_files = []
    for root, dirs, files in os.walk(selected_folder):
        for file in files:
            if pattern in file:
                full_path = os.path.join(root, file)
                matching_files.append(full_path)
                
    return matching_files


#####################################################################################################################################
#           Extraction de la base au format EXCEL                                                                                   #
#####################################################################################################################################  
def db_to_excel(conn, excel_filename):
    """
    Extrait les données des tables CAVITE, ENTREE et DOCUMENT d'une base SQLite
    et les enregistre dans un fichier Excel, avec les champs de ENTREE et DOCUMENT numérotés,
    ordonnés logiquement après les champs de CAVITE.

    :param conn: Connexion sqlite3.Connection ouverte
    :param excel_filename: Nom du fichier Excel de sortie
    :param max_entries: Nombre max de ENTREE par cavité à inclure
    :param max_documents: Nombre max de DOCUMENT par cavité à inclure
    """
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT MAX(doc_count) 
        FROM (
            SELECT COUNT(*) AS doc_count
            FROM DOCUMENT
            GROUP BY ID_CAVITE
        )
    """)
    result = cursor.fetchone()
    max_documents = result[0] if result[0] is not None else 0

    # query = """
    #     SELECT *
    #     FROM ENTREE
    #     INNER JOIN CAVITE ON CAVITE.ID = ENTREE.ID_CAVITE
    # """
    
    # df_cavite_2 = pd.read_sql_query(query, conn)

    # Charger les tables
    df_cav = pd.read_sql_query("SELECT * FROM CAVITE", conn)
    df_cav.columns = [f"CAVITE_{col}" for col in df_cav.columns]

    df_ent = pd.read_sql_query("SELECT * FROM ENTREE", conn)
    df_ent.columns = [f"ENT_{col}" for col in df_ent.columns]
       
    df_cavite = df_ent.merge(df_cav, how='left', left_on='ENT_ID_CAVITE', right_on='CAVITE_ID')

    df_document = pd.read_sql_query("SELECT * FROM DOCUMENT", conn)
    
    # Idem pour DOCUMENT
    df_document['DOCUMENT_NUM'] = df_document.groupby('ID_CAVITE').cumcount() + 1
    df_document = df_document[df_document['DOCUMENT_NUM'] <= max_documents]

    df_document_flat = df_document.pivot(index='ID_CAVITE', columns='DOCUMENT_NUM')
    df_document_flat.columns = [f"DOCUMENT_{num}_{col}" for col, num in df_document_flat.columns]
    df_document_flat.reset_index(inplace=True)
    
    # print(df_document_flat)
    
    colonnes_tries = sorted(df_document_flat.columns)
    # document_columns = sorted([col for col in df_document_flat.columns if col.startswith('DOCUMENT_')], key=lambda x: (int(x.split('_')[1]), x))
  
    df_document_flat = df_document_flat[colonnes_tries]
 
    # print(df_document_flat)

    # Fusion avec CAVITE
    df_cavite['_est_premiere'] = ~df_cavite.duplicated(subset='ENT_ID_CAVITE', keep='first')

    # Joindre normalement    
    df_merged = df_cavite.merge(df_document_flat, how='left',  left_on='ENT_ID_CAVITE', right_on='ID_CAVITE')

    # Pour les lignes qui ne sont pas les premières : vider les colonnes venant de df_document_flat
    colonnes_document = [col for col in df_document_flat.columns if col != 'ID_CAVITE']
    df_merged.loc[~df_merged['_est_premiere'], colonnes_document] = pd.NA

    # Nettoyage
    df = df_merged.drop(columns=['_est_premiere'])
    
    # print(df)

    # Supprimer colonnes de jointure inutiles
    df.drop(columns=['ID_CAVITE'], inplace=True, errors='ignore')
    df.drop(columns=[col for col in df.columns if '_TH_VALIDE' in col], inplace=True, errors='ignore')

   
    # Export Excel
    df.to_excel(excel_filename, index=False)
         
    log.info(f"Exportation excel terminée : {Colors.WHITE}{excel_filename}")


#####################################################################################################################################
#           Mise en forme du fichier EXCEL                                                                                          #
#####################################################################################################################################  
def adapt_excel(excel_filename, selected_folder, update, max_documents=5):
    
    # Chargement pour édition
    wb = load_workbook(excel_filename)
    ws = wb.active

    # Définition des styles
    yellow_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")  # Jaune moyen
    yellow_low_fill = PatternFill(start_color="FFEA8F", end_color="FFEA8F", fill_type="solid")  # Jaune moyen
    green_fill = PatternFill(start_color="99CC00", end_color="99CC00", fill_type="solid")   # Vert clair
    green_low_fill = PatternFill(start_color="E5FF9B", end_color="E5FF9B", fill_type="solid")   # Vert clair
    blue_fill = PatternFill(start_color="2F97FF", end_color="2F97FF", fill_type="solid")   # bleu 
    blue_low_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")   # bleu 
    gris_fill = PatternFill(start_color="777777", end_color="777777", fill_type="solid")  # Style de fond gris clair
    gris_low_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Style de fond gris clair
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # Style de fond gris clair
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) # Définir le style pour l'en-tête
    header_font = Font(bold=True) # Définir le style pour l'en-tête



    # Identifier les colonnes à colorier selon leur nom en première ligne
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value is None:
            continue
        if str(header_value).startswith("ENT_"):
            fill = yellow_low_fill
        elif str(header_value).startswith("CAVITE_"):
            fill = green_low_fill
        elif str(header_value).startswith("DOCUMENT_"):
            fill = blue_low_fill
        else:
            continue

        # Appliquer la couleur aux cellules de la colonne (sauf entête si tu veux)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.fill = fill
                
        
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value is None:
            continue
        if str(header_value).endswith("_ID"):
            fill = gris_fill
        elif str(header_value).endswith("_ID_CAVITE"):
            fill = gris_fill
        elif str(header_value).endswith("_PATH"):
            fill = gris_fill
        elif str(header_value).endswith("_KEY_KARSTEAU"):
            fill = gris_low_fill    
        elif str(header_value).endswith("_DATE_UPDATE"):
            fill = gris_fill
        elif str(header_value).endswith("_HASH_SQL_FILE"):
            fill = gris_fill
        elif str(header_value).endswith("_HASH_FILE"):
            fill = gris_fill          
        else:
            continue

        # Appliquer la couleur aux cellules de la colonne (sauf entête si tu veux)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.fill = fill


    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 2  # +2 pour un peu de marge
        
    # Définir une bordure fine sur tous les côtés
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Appliquer la bordure à toutes les cellules sauf l'en-tête
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Appliquer le style à chaque cellule de la première ligne
    for cell in ws[1]:
        cell.alignment = header_alignment
        cell.font = header_font
        cell.fill = header_fill

    # Définir la hauteur de ligne à 40 pour la ligne 1
    ws.row_dimensions[1].height = 40
    
    # Insérer une nouvelle ligne en haut
    ws.insert_rows(1)

    # Mettre un message dans A1
    ws['A1'].value = f"Tableau d'échange Therion - Karsteau, base : [{os.path.basename(selected_folder)}], mise à jour le : {update}"

    # Appliquer le style à toutes les cellules de la nouvelle première ligne
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.font = header_font
        cell.fill = header_fill


    # Hauteur de ligne
    ws.row_dimensions[1].height = 40

    # Sauvegarder les modifications
    wb.save(excel_filename)
    
    log.info(f"Mise en forme fichier excel terminée : {Colors.WHITE}{excel_filename}")



#####################################################################################################################################
#                                                                                                                                   #
#                                                           Main                                                                    #
#                                                                                                                                   #
#####################################################################################################################################
if __name__ == '__main__':
    error_count = 0
    warning_fix = 0
    outputs_path = "./Test/"
    inputs_path = "./Test/"

    maintenant = datetime.now()
    
    parser = argparse.ArgumentParser(description="Export d'une base de données Therion. ", 
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--folder", help="Chemin vers le dossier de la bd à exporter")
    parser.epilog = (f"Sélectionner le dossier source de le base de données Therion")
   
    # Analyser les arguments de ligne de commande
    args = parser.parse_args()

    if not args.folder:    # Si aucun fichier n'est fourni en ligne de commande, ouvrir une fenêtre Tkinter pour sélectionner un fichier
        if os.name == 'posix':  os.system('clear') # Linux, MacOS
        elif os.name == 'nt':  os.system('cls')# Windows
        else: print("\n" * 100)
        
        root = tk.Tk()
        root.withdraw()  # Cacher la fenêtre principale de Tkinter
        selected_folder = filedialog.askdirectory(title="Sélectionnez le dossier de la bd Therion")
        
        # selected_folder =  "C:/Users/alexa/Desktop/pyThtoBD/Test/TKM-Test"
        
        outputs_path = os.path.dirname(selected_folder) + "/"
           
    else :         # Si le fichier est fourni en ligne de commande 
        input_folder = args.folder
        if os.path.isdir(input_folder) is False :
            print(f"{Colors.ERROR}Erreur : le dossier {Colors.CYAN}{input_folder}{Colors.ERROR} est inexistant")      
            print(f"{Colors.GREEN}Commande : {Colors.WHITE}python pyThtoBD.py --folder ./chemin/dossier/")
            sys.exit()  
        else :
            selected_folder = os.path.abspath(input_folder)
            if os.name == 'posix':  os.system('clear') # Linux, MacOS
            elif os.name == 'nt':  os.system('cls')# Windows
            else: print("\n" * 100)
        
    output_folder = selected_folder + export_folder
    output_db = selected_folder + export_folder + export_db
    update = maintenant.strftime("%Y-%m-%d %H:%M:%S")
      
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"Dossier '{output_folder}' créé.")
    else:
        print(f"Dossier '{output_folder}' existe déjà.")
    
    log = setup_logger(output_folder + log_file, debug_log)
    
    # log.debug("Ceci est un message de debug")
    # log.info("Tout va bien")
    # log.warning("Attention, possible souci")
    # log.error("Une erreur est survenue")
    # log.critical("Erreur critique !")
       
    _titre =[f'********************************************************************************************************************************************\033[0m', 
             f'* Export d\'une BD Therion',
             f'*       Script pyThtoBD par :{Colors.MAGENTA} alexandre.pont@yahoo.fr',
             f'*       Version : {Colors.MAGENTA}' + Version,
             f'*       Dossier source : {Colors.MAGENTA}' + selected_folder,           
             f'*       Dossier destination : {Colors.MAGENTA}' + output_folder,
             f'*       Base de données destination : {Colors.MAGENTA}' + output_db,
             f'*       Date : {Colors.MAGENTA}' + update,
             f'*       ',
             f'*       ',
             f'********************************************************************************************************************************************\033[0m']     
    
    for i in range(11): log.info(_titre[i])
    
    if os.path.isfile(output_db) is False :
        log.info("Le fichier : " + output_db + " n’existe pas, création automatique") 
        conn = create_new_db(output_db)
        
    else : 
        log.debug("Le fichier : " + output_db + " existe déjà, ouverture pour mise à jour")
        conn = sqlite3.connect(output_db)
        init_th_valide(conn)
        
    file_list = find_files(selected_folder, export_file)
    log.info(f"Nombre de cavités à traiter : {Colors.MAGENTA}{len(file_list)}")
        
    cavite_update(conn, file_list, selected_folder, update)
    
    output  = output_db[:-3]
    zip_file(selected_folder, selected_folder  + export_folder, output)  
    
    output  = output_db[:-3] + "_data.zip"
    zip_data(conn, selected_folder  + export_folder, selected_folder, output) 
    
    output  = output_db[:-3] + ".xlsx"
    db_to_excel(conn, output)
    
    adapt_excel(output, selected_folder, update)
    
    conn.close()
    
    if warning_fix > 0 :
         log.warning(f"""Nbre de point(s) fixe trouvé(s) : {Colors.CYAN}{warning_fix}{Colors.YELLOW}, vérifier : une entrée doit avoir l'attribut type {Colors.CYAN}station 0 "Entrée XXXX" entrance{Colors.YELLOW} lors de sa déclaration""") 
    
    if error_count > 0 :
         log.error(f"""Nbre d'erreur(s)  trouvé(s) : {Colors.CYAN}{error_count}{Colors.YELLOW}, à vérifier""")
    else :         
        log.info("Aucune d'erreur trouvée, parfait !")

